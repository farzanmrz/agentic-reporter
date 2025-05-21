# Import self funcs
import csv
import importlib
import itertools
import re
import sys
import warnings
from datetime import datetime, time
from typing import Optional, Tuple

import openpyxl

# Import general usecase libraries
import pandas as pd
import torch

# Import parsing libraries
import xlrd
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from xlrd.xldate import XLDateAmbiguous, XLDateError, xldate_as_tuple


# [0a] Returns lowercase without spaces list of strings
def normalize_string(cell_text):
    return str(cell_text).lower().replace(" ", "") if cell_text is not None else ""


# [1a] Datatype validation function
def xlsx_dataType(value, number_format):
    """
    Determines the type of the cell and handles necessary conversions.

    Int to Type Key Mapping:
        0  -> Empty cell
        1  -> Text cell
        2  -> Numeric type
        3  -> Integer subclass of number
        4  -> Float subclass of number
        5  -> Percentage subclass of number
        6  -> Currency subclass of number
        7  -> Scientific subclass of number
        8  -> Date type
        9  -> Time subclass of date
        10 -> Datetime subclass of date
        11 -> Boolean type with T or F value
        12 -> Error type corresponding to #REF!, #VALUE! etc. in excel
        13 -> Blank cell with formatting
        14 -> Unknown type not in our keys

    Args:
        value: The value of the cell.
        number_format (str): The number format of the cell.

    Returns:
        int: The determined type of the cell (0-14).
    """
    # Empty cell with no value return blank for this
    if value is None:
        return 13

    # Boolean cell
    elif isinstance(value, bool):
        return 11

    # Error cell based on certain error values in Excel
    elif any(
        error in str(value)
        for error in {
            "#VALUE!",
            "#REF!",
            "#DIV/0!",
            "#NAME?",
            "#N/A",
            "#NUM!",
            "#NULL!",
        }
    ):
        return 12

    # Currency format cell
    elif "#,##0" in number_format:
        return 6

    # Percentage format cell
    elif number_format.endswith("%"):
        return 5

    # Check for scientific notation format
    elif "E+" in number_format or "E-" in number_format:
        return 7

    # If float type
    elif isinstance(value, float):
        # Check if it's an integer subclass
        if value.is_integer():
            return 3
        # Otherwise, it's a float
        return 4

    # If integer type
    elif isinstance(value, int):
        return 3

    # Time type cell
    elif isinstance(value, time):
        return 9

    # Datetime type cell
    elif isinstance(value, datetime):
        # Check if the time part is exactly midnight
        if value.time() == time(0, 0):
            # If the datetime is formatted as a date, return it as a date type
            if "d" in number_format.lower() and "h" not in number_format.lower():
                return 8
            else:
                # If there is significant time information, return it as datetime
                return 10
        else:
            # If the time part is not midnight, it's definitely a datetime
            return 10

    # Blank cell with formatting
    elif value == "":
        return 13

    # Text cell
    elif isinstance(value, str):
        return 1

    # Return 14 for any unknown type
    return 14


# [1b] Get the parsed value of data in cell
def xlsx_content(cell_type, cell):
    # If empty cell then return an empty string
    if cell_type == 0:
        return ""

    # Text cell
    elif cell_type == 1:
        return str(cell.value)

    # Numeric cell default type
    elif cell_type == 2:
        return str(cell.value)

    # Integer subclass of number, remove proceeding 0s and return str
    elif cell_type == 3:
        return str(cell.value)

    # Decimal subclass then convert to float and return str
    elif cell_type == 4:
        return str(cell.value)

    # Percentage subclass then convert to value add % sign and return the string
    elif cell_type == 5:
        return str(cell.value * 100) + "%"

    # Currency subclass then convert to value add currency symbol and return the string
    elif cell_type == 6:
        # Check if cell value already is a string so it will have the symbol
        if isinstance(cell.value, str):
            return str(cell.value)
        # Check if euro sign in format string because that is separate
        if "$€" in cell.number_format:
            return "€" + str(float(cell.value))
        # In other case for usd values we gotta add the $ symbol
        return "$" + str(float(cell.value))

    # If scientific subclass then convert to value add 'E' sign and return the string
    elif cell_type == 7:
        if isinstance(cell.value, (int, float)):
            return f"{cell.value:.2e}"
        else:
            return str(cell.value)

    # If date subclass then use the df to return value
    elif cell_type == 8:
        return str(cell.value).split(" ")[0]

    elif cell_type in [9, 10]:
        return str(cell.value)

    # If boolean cell then use df to get string
    elif cell_type == 11:
        return str(cell.value)

    # If error cell then use df to get string
    elif cell_type == 12:
        return str(cell.value)

    # If blank cell (Empty cell with formatting) then return an empty string
    elif cell_type == 13:
        return ""

    # Return empty string for default case
    return ""


# [2] Main Verification Function definition
def verify_xlsx(
    file_path, SHEET_NUMBER=1, HEADER_ROW_IDX=0, MONITOR_COLS=[], REPORT_COLS=[]
):

    # Suppress all UserWarnings related to various xlsx issues
    warnings.filterwarnings("ignore", category=UserWarning)

    """Step 1. Try loading the workbook"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

    # If file not found, return an error message
    except FileNotFoundError:
        return {"status": "error", "message": f"XLSX File not found: {file_path}"}

    # Catch other openpyxl loading errors
    except Exception as e:
        return {
            "status": "error",
            "message": f"Error loading XLSX workbook '{file_path}': {str(e)}",
        }

    """Step 2. Verify SHEET_NUMBER is in valid range"""
    num_sheets = len(workbook.worksheets)
    if not (1 <= SHEET_NUMBER <= len(workbook.worksheets)):
        return {
            "status": "error",
            "message": f"SHEET_NUMBER ({SHEET_NUMBER}) is invalid. The workbook has {num_sheets} sheet(s). Please enter a number between 1 and {num_sheets}.",
        }

    """Step 3. Check if the sheet exists and error if not"""
    try:
        target_sheet = workbook.worksheets[SHEET_NUMBER - 1]
    except Exception as e:
        return {
            "status": "error",
            "message": f"Internal error: Could not access sheet at index {actual_sheet_index}.",
        }

    """Step 4. Acess the sheet and get the header row"""

    # Create normalized tracking sets for faster lookups
    monitor_check = {normalize_string(col): col for col in MONITOR_COLS}
    report_check = {normalize_string(col): col for col in REPORT_COLS}

    # Define list to store the extracted column names
    header_row_cells = []
    header_cell_idxs = []

    # Loop through all the columns in the target sheet
    for col in range(target_sheet.max_column):

        # 4a. Extract the cell from the HEADER_ROW_IDX row and the current column
        header_cell = target_sheet.cell(
            row=HEADER_ROW_IDX + 1, column=col + 1
        )  # SIde note check what to do if this is empty or None

        # 4b. Get the datatype of the cell
        header_cell_datatype = xlsx_dataType(
            header_cell.value, header_cell.number_format
        )

        # 4c. Get the actual text/content of the cell parsed according to type
        header_cell_text = xlsx_content(header_cell_datatype, header_cell)

        # 4d. Normalize the text
        norm_text = normalize_string(header_cell_text)

        # 4e. Check if current column header is a monitor/report col
        if norm_text in monitor_check:

            # If it is then append the text and col index
            header_row_cells.append(header_cell_text)
            header_cell_idxs.append(col)
            del monitor_check[norm_text]
        elif norm_text in report_check:
            header_row_cells.append(header_cell_text)
            header_cell_idxs.append(col)
            del report_check[norm_text]

    """Step 5. Check what to return finally"""
    if monitor_check or report_check:
        return {
            "status": "error",
            "missing_monitor_cols": list(monitor_check.values()),
            "missing_report_cols": list(report_check.values()),
        }

    # If everything matched successfully
    return {
        "status": "success",
        "header_cols": header_row_cells,
        "header_col_idxs": header_cell_idxs,
    }
